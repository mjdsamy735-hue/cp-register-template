"""Generate a CP item assist register workbook from TRU source data.

The script accepts a CSV file with TRU/item information and produces an Excel
workbook with a formatted CP Register sheet, a TRU Source Data sheet, and a
lookup sheet for controlled values.
"""

from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

REGISTER_HEADERS = [
    "CP Item ID",
    "Discipline",
    "System",
    "Subsystem",
    "TRU Tag",
    "TRU Description",
    "CP Activity",
    "Location",
    "Priority",
    "Responsible Party",
    "Status",
    "Required Evidence",
    "Source Reference",
    "Comments",
]

SOURCE_HEADERS = [
    "tru_tag",
    "tru_description",
    "discipline",
    "system",
    "subsystem",
    "location",
    "priority",
    "responsible_party",
    "source_reference",
    "comments",
]

STATUS_VALUES = ["Not Started", "In Progress", "Ready for Review", "Closed", "On Hold"]
PRIORITY_VALUES = ["High", "Medium", "Low"]


@dataclass(frozen=True)
class TruRecord:
    tru_tag: str
    tru_description: str
    discipline: str = ""
    system: str = ""
    subsystem: str = ""
    location: str = ""
    priority: str = "Medium"
    responsible_party: str = ""
    source_reference: str = ""
    comments: str = ""

    @classmethod
    def from_row(cls, row: dict[str, str]) -> "TruRecord":
        values = {key: (row.get(key) or "").strip() for key in SOURCE_HEADERS}
        if not values["tru_tag"]:
            raise ValueError("Every source row must include a tru_tag value.")
        if values["priority"] and values["priority"] not in PRIORITY_VALUES:
            raise ValueError(
                f"Invalid priority for {values['tru_tag']!r}: {values['priority']!r}. "
                f"Expected one of: {', '.join(PRIORITY_VALUES)}"
            )
        if not values["priority"]:
            values["priority"] = "Medium"
        return cls(**values)


def read_tru_records(source_csv: Path) -> list[TruRecord]:
    with source_csv.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise ValueError("Input CSV is empty or missing headers.")

        missing = sorted(set(["tru_tag", "tru_description"]) - set(reader.fieldnames))
        if missing:
            raise ValueError(f"Input CSV is missing required columns: {', '.join(missing)}")

        return [TruRecord.from_row(row) for row in reader]


def build_cp_activity(record: TruRecord) -> str:
    if record.tru_description:
        return f"Verify and register CP requirements for {record.tru_description}"
    return "Verify and register CP requirements"


def build_evidence(record: TruRecord) -> str:
    details = ["Completed CP checklist"]
    if record.system:
        details.append(f"system evidence for {record.system}")
    if record.tru_tag:
        details.append(f"TRU reference {record.tru_tag}")
    return "; ".join(details)


def autosize_columns(ws, min_width: int = 12, max_width: int = 42) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_letter].width = min(max(length + 2, min_width), max_width)


def add_table(ws, name: str, last_row: int, last_col: int) -> None:
    if last_row < 2:
        return
    ref = f"A1:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    autosize_columns(ws)


def write_lookup_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Lookups")
    ws.append(["Status", "Priority"])
    max_rows = max(len(STATUS_VALUES), len(PRIORITY_VALUES))
    for index in range(max_rows):
        ws.append([
            STATUS_VALUES[index] if index < len(STATUS_VALUES) else "",
            PRIORITY_VALUES[index] if index < len(PRIORITY_VALUES) else "",
        ])
    style_sheet(ws)
    ws.sheet_state = "hidden"


def add_validations(ws, max_row: int) -> None:
    status_validation = DataValidation(type="list", formula1='"Not Started,In Progress,Ready for Review,Closed,On Hold"')
    priority_validation = DataValidation(type="list", formula1='"High,Medium,Low"')
    ws.add_data_validation(status_validation)
    ws.add_data_validation(priority_validation)

    priority_col = REGISTER_HEADERS.index("Priority") + 1
    status_col = REGISTER_HEADERS.index("Status") + 1
    priority_validation.add(f"{get_column_letter(priority_col)}2:{get_column_letter(priority_col)}{max_row}")
    status_validation.add(f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}{max_row}")


def write_register_sheet(wb: Workbook, records: Iterable[TruRecord]) -> None:
    ws = wb.active
    ws.title = "CP Register"
    ws.append(REGISTER_HEADERS)

    for index, record in enumerate(records, start=1):
        ws.append([
            f"CP-{index:04d}",
            record.discipline,
            record.system,
            record.subsystem,
            record.tru_tag,
            record.tru_description,
            build_cp_activity(record),
            record.location,
            record.priority,
            record.responsible_party,
            "Not Started",
            build_evidence(record),
            record.source_reference,
            record.comments,
        ])

    style_sheet(ws)
    add_validations(ws, max(ws.max_row, 2))
    add_table(ws, "CPRegister", ws.max_row, len(REGISTER_HEADERS))


def write_source_sheet(wb: Workbook, records: Iterable[TruRecord]) -> None:
    ws = wb.create_sheet("TRU Source Data")
    ws.append(SOURCE_HEADERS)
    for record in records:
        ws.append([getattr(record, header) for header in SOURCE_HEADERS])
    style_sheet(ws)
    add_table(ws, "TRUSourceData", ws.max_row, len(SOURCE_HEADERS))


def create_workbook(records: list[TruRecord], output_path: Path) -> None:
    if not records:
        raise ValueError("No TRU rows found in the input CSV.")

    wb = Workbook()
    write_register_sheet(wb, records)
    write_source_sheet(wb, records)
    write_lookup_sheet(wb)
    wb.properties.title = "CP Register Template"
    wb.properties.subject = "CP item assist register generated from TRU data"
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create a CP register Excel workbook from TRU CSV data.")
    parser.add_argument("input_csv", type=Path, help="Path to TRU source CSV file.")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path("cp_register_template.xlsx"),
        help="Output Excel workbook path. Defaults to cp_register_template.xlsx.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    records = read_tru_records(args.input_csv)
    create_workbook(records, args.output)
    print(f"Created {args.output} with {len(records)} CP register rows.")


if __name__ == "__main__":
    main()
